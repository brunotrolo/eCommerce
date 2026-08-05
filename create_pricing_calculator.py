#!/usr/bin/env python3
"""
Shopee Pricing Calculator - Generates Excel calculator with formulas
Based on actual Shopee retention rates from order analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

def create_pricing_calculator():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculadora"

    # Define colors and styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    input_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    result_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40

    # Title
    ws['A1'] = "CALCULADORA DE PREÇOS SHOPEE"
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color="1F4E78")
    ws.merge_cells('A1:C1')

    ws['A2'] = "Baseada em análise de 5 vendas reais com taxas médias da Shopee"
    ws['A2'].font = Font(name='Arial', size=9, italic=True)
    ws.merge_cells('A2:C2')

    # Section 1: Inputs
    ws['A4'] = "ENTRADA DE DADOS"
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws.merge_cells('A4:C4')

    row = 5
    ws[f'A{row}'] = "Preço de Compra (Custo)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "R$ "
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '#,##0.00'

    row = 6
    ws[f'A{row}'] = "Taxa de Retenção Shopee (%)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "Selecione a taxa baseada no cenário"
    ws[f'C{row}'].font = Font(size=9, italic=True)

    row = 7
    ws[f'A{row}'] = "  ○ Sem promoção (comissão pura)"
    ws[f'B{row}'] = 24.5
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '0.0'

    row = 8
    ws[f'A{row}'] = "  ○ Com Pix de desconto (típico)"
    ws[f'B{row}'] = 28.5
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '0.0'

    row = 9
    ws[f'A{row}'] = "  ○ Com voucher + Pix (máximo)"
    ws[f'B{row}'] = 32.0
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '0.0'

    row = 10
    ws[f'A{row}'] = "Taxa Selecionada para Cálculo"
    ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    ws[f'B{row}'] = "=B8"
    ws[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '0.0'
    ws[f'C{row}'] = "Edite para usar outra taxa"
    ws[f'C{row}'].font = Font(size=9, italic=True)

    row = 12
    ws[f'A{row}'] = "Margem de Lucro Desejada (%)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 15.0
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '0.0'
    ws[f'C{row}'] = "Padrão: 15% (ajuste conforme necessário)"

    # Section 2: Calculations
    ws['A14'] = "CÁLCULOS"
    ws['A14'].font = header_font
    ws['A14'].fill = header_fill
    ws.merge_cells('A14:C14')

    row = 15
    ws[f'A{row}'] = "Preço Mínimo (Break-even)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "= Custo ÷ (1 - Taxa de Retenção)"
    ws[f'C{row}'].font = Font(size=9, italic=True)
    ws[f'B{row}'] = "=B5/(1-B10/100)"
    ws[f'B{row}'].fill = result_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'R$ #,##0.00'
    ws[f'B{row}'].font = Font(bold=True)

    row = 16
    ws[f'A{row}'] = "Preço Sugerido (com Margem)"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "= Break-even × (1 + Margem)"
    ws[f'C{row}'].font = Font(size=9, italic=True)
    ws[f'B{row}'] = f"=B15*(1+B12/100)"
    ws[f'B{row}'].fill = result_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'R$ #,##0.00'
    ws[f'B{row}'].font = Font(bold=True, size=12, color="006100")

    row = 18
    ws[f'A{row}'] = "Análise de Rentabilidade"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells('A18:C18')

    row = 19
    ws[f'A{row}'] = "Receita Bruta (Preço Sugerido)"
    ws[f'B{row}'] = "=B16"
    ws[f'B{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'R$ #,##0.00'

    row = 20
    ws[f'A{row}'] = "Retenção Shopee (em R$)"
    ws[f'B{row}'] = "=B16*(B10/100)"
    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'R$ #,##0.00'

    row = 21
    ws[f'A{row}'] = "Receita Líquida (Shopee)"
    ws[f'B{row}'] = "=B16-B20"
    ws[f'B{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'R$ #,##0.00'
    ws[f'B{row}'].font = Font(bold=True)

    row = 22
    ws[f'A{row}'] = "Custo do Produto"
    ws[f'B{row}'] = "=B5"
    ws[f'B{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'R$ #,##0.00'

    row = 23
    ws[f'A{row}'] = "LUCRO LÍQUIDO POR VENDA"
    ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
    ws[f'B{row}'] = "=B21-B22"
    ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'R$ #,##0.00'
    ws[f'B{row}'].font = Font(bold=True, size=12, color="006100")

    row = 24
    ws[f'A{row}'] = "Margem Líquida (%)"
    ws[f'B{row}'] = "=IF(B21=0,0,(B23/B21)*100)"
    ws[f'B{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '0.00"%"'

    # Section 3: Cenários
    ws['A26'] = "SIMULADOR DE CENÁRIOS"
    ws['A26'].font = header_font
    ws['A26'].fill = header_fill
    ws.merge_cells('A26:C26')

    row = 27
    headers = ["Custo", "Taxa", "Preço Mínimo", "Preço Sugerido", "Lucro/unidade"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Example scenarios
    scenarios = [
        ("Exemplo: R$50", 50, 24.5),
        ("Exemplo: R$100", 100, 24.5),
        ("Exemplo: R$200", 200, 28.5),
    ]

    for idx, (label, cost, rate) in enumerate(scenarios, 28):
        ws[f'A{idx}'] = label
        ws[f'A{idx}'].border = thin_border
        ws[f'B{idx}'] = cost
        ws[f'B{idx}'].border = thin_border
        ws[f'B{idx}'].number_format = 'R$ #,##0.00'
        ws[f'C{idx}'] = rate
        ws[f'C{idx}'].border = thin_border
        ws[f'C{idx}'].number_format = '0.0'
        ws[f'D{idx}'] = f"=B{idx}/(1-C{idx}/100)"
        ws[f'D{idx}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws[f'D{idx}'].border = thin_border
        ws[f'D{idx}'].number_format = 'R$ #,##0.00'
        ws[f'E{idx}'] = f"=D{idx}*1.15"
        ws[f'E{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws[f'E{idx}'].border = thin_border
        ws[f'E{idx}'].number_format = 'R$ #,##0.00'
        ws[f'F{idx}'] = f"=(E{idx}*(1-C{idx}/100))-B{idx}"
        ws[f'F{idx}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws[f'F{idx}'].border = thin_border
        ws[f'F{idx}'].number_format = 'R$ #,##0.00'

    # Notes section
    ws['A32'] = "NOTAS E EXPLICAÇÕES"
    ws['A32'].font = header_font
    ws['A32'].fill = header_fill
    ws.merge_cells('A32:C32')

    notes = [
        "1. Taxa de Retenção: Total de comissões, taxas de serviço, descontos Pix e rebates que Shopee retém",
        "2. Preço Mínimo: Preço necessário apenas para cobrir custos (sem lucro)",
        "3. Preço Sugerido: Preço Mínimo + 15% de margem (ajustável conforme estratégia)",
        "4. Receita Líquida: O valor que você realmente recebe após taxas Shopee",
        "5. Lucro: Receita Líquida - Custo do Produto",
        "",
        "TABELA DE RETENÇÃO OBSERVADA NAS 5 VENDAS ANALISADAS:",
        "• Sem promoção: 24.5% (pedido com comissão simples)",
        "• Com Pix de desconto: 25.9% a 32.4% (depende de rebate e coins)",
        "• Caso típico: 28.5% (recomendado para novos produtos)",
        "",
        "DICA: Para produtos novos, use 28.5%. Conforme a promoção progride,",
        "ajuste conforme a taxa observada em seus pedidos reais."
    ]

    for idx, note in enumerate(notes, 33):
        ws[f'A{idx}'] = note
        ws[f'A{idx}'].font = Font(size=9)
        ws.merge_cells(f'A{idx}:C{idx}')

    # Freeze panes
    ws.freeze_panes = 'A5'

    wb.save('/home/user/eCommerce/Calculadora_Precos_Shopee.xlsx')
    print("✓ Calculadora criada: Calculadora_Precos_Shopee.xlsx")

if __name__ == "__main__":
    create_pricing_calculator()
